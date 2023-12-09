from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
import time
from tkcalendar import DateEntry
from datetime import date, datetime, timedelta  # Importar timedelta
from tkinter import messagebox
import sqlite3
import pandas as pd
import locale
from babel import numbers
import sys
import os
import configparser
import socket
from tkinter import Tk, Toplevel, Button
from tkinter import filedialog

import openpyxl
from openpyxl.styles import PatternFill, Font


# funcion para desabilitar los botones (supervisores, ingresar, y cancelar )
def deshabilitar_botones():
    botones = [btn_ingresar, castellar, gomez,
               Castrillo, Virgilio,  boton_salida, gomezJose]
    cancelar_etiquetas = [cancelar1, cancelar2,
                          cancelar3, cancelar4, cancelar5]

    for boton in botones:
        boton.configure(state="disabled")

    for etiqueta in cancelar_etiquetas:
        etiqueta.configure(state="disabled")

# funcion para habilitar los botones (supervisores, ingresar)


def habilitar_botones():
    botones = [castellar, gomez, Castrillo, Virgilio, gomezJose]
    cancelar_etiquetas = [cancelar1, cancelar2,
                          cancelar3, cancelar4, cancelar5]

    btn_ingresar.configure(state="normal")
    boton_salida.configure(state="normal")

    for boton, cancelar in zip(botones, cancelar_etiquetas):
        boton.configure(state="normal")

        if boton["bg"] == "red":
            cancelar.configure(state="normal")
        else:
            cancelar.configure(state="disabled")

# Función para animar el label


def animar_label(label):
    label["fg"] = "red2"
    ventana.after(600, lambda: label.config(bg=ventana.cget("bg"), fg="black"))


def ani_label_5min(label):
    label["fg"] = "DarkOrange2"
    ventana.after(600, lambda: label.config(bg=ventana.cget("bg"), fg="black"))

# Crear botón Castellar, cancelar y cronómetro


def estilos_botones_supervisores(Btn):
    Btn.configure(
        bg="green2",
        fg="black",
        activebackground="red",
        activeforeground="white",
        cursor="hand2",
        font=("Arial", 12, "bold"),
        width=8,
        height=1
    )


def det_cancelar(cancelar):
    cancelar.config(
        image=cancelar_img,
        cursor="hand2",
        relief="groove"
    )


def detalledateEntry(date_Entry):
    date_Entry.configure(
        font=("Segoe UI", 12),
        justify='center',
        background='#3F3D56'
    )

# estilos de los widgets


def establecer_estilos(widget):
    widget.configure(
        bg="#26FF00",
        fg="black",
        activebackground="#26FF00",
        activeforeground="black",
        font=("Arial", 12, "bold")
    )

# funcion para botones segunda pantalla


def boton_clave(widget):
    widget.configure(
        bg="#06B1AB",
        cursor="hand2",
        fg="black",
        activebackground="#06B1AB",
        activeforeground="black",
        font=("Arial", 12, "bold"),
        compound="left",
        padx=3,
        width=87,
        height=20
    )

# funcion para botones segunda pantalla


def botones_ingresar(widget):
    widget.configure(
        bg="#06B1AB",
        cursor="hand2",
        fg="black",
        activebackground="#06B1AB",
        activeforeground="black",
        font=("Arial", 12, "bold"),
        compound="left",
        padx=3,
        width=87,
        height=20
    )

# funcion para botones Salidas


def estilo_botones_Salidas(widget):
    widget.configure(
        cursor="hand2",
        bg="red",
        fg="white",
        activebackground="red",
        activeforeground="white",
        font=("Arial", 12, "bold"),
        compound="left",
        padx=5,
        width=70,
        height=20
    )

# Funciones para mover la ventana principal (igual que en el código anterior)


def guardar_posi_principal(event):
    global x_origen, y_origen
    x_origen, y_origen = event.x, event.y


def mover_ventana_principal(event):
    x, y = ventana.winfo_x() + event.x - x_origen, ventana.winfo_y() + event.y - y_origen
    if x < 0:
        x = 0
    if y < 0:
        y = 0
    ventana.geometry(f"+{x}+{y}")

# Función para obtener la posición actual de la ventana principal con respecto a la pantalla


def obtener_posicion_ventana_principal():
    x = ventana.winfo_rootx()
    y = ventana.winfo_rooty()
    return x, y


# Leer la configuración desde el archivo .ini
config = configparser.ConfigParser()
config.read('config.ini')

# Obtener la ruta de la base de datos y el nombre de la terminal desde la configuración
ruta_basedatos = config.get('Configuracion', 'ruta_basedatos')
nombre_terminal = config.get('Configuracion', 'nombre_terminal')


# Crear ventana principal
ventana = Tk()
ventana.geometry("260x360")
ventana.title("Supervisores")
ventana.wm_attributes("-topmost", True)
ventana.overrideredirect(True)
ventana.configure(bg="gray88")

ventana.bind("<Button-1>", guardar_posi_principal)
ventana.bind("<B1-Motion>", mover_ventana_principal)

# Funcion para cargar las imagenes


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# Cargar Imagenes
ruta_salida = resource_path("Iconos/salida.png")
Salir_img = ImageTk.PhotoImage(Image.open(ruta_salida))

ruta_regresar = resource_path("Iconos/back.png")
Regresar_img = ImageTk.PhotoImage(Image.open(ruta_regresar))

ruta_open = resource_path("Iconos/accept_1.png")
open_img = ImageTk.PhotoImage(Image.open(ruta_open))

ruta_export = resource_path("Iconos/save.png")
Export_img = ImageTk.PhotoImage(Image.open(ruta_export))

ruta_cancelar = resource_path("Iconos/cancelado.png")
cancelar_img = ImageTk.PhotoImage(Image.open(ruta_cancelar))

ruta_ingresar = resource_path("Iconos/next.png")
Ingresar_img = ImageTk.PhotoImage(Image.open(ruta_ingresar))

ruta_reloj = resource_path("Iconos/reloj-32.png")
Reloj_img = ImageTk.PhotoImage(Image.open(ruta_reloj))

# Pantalla Home
def home():
    home_label.config(
        fg="white",
        bg="black",
        font=("Segoe UI", 11),
        justify="center"
    )

    # Configurar la columna 0 y la columna 2 para que se expandan horizontalmente
    ventana.columnconfigure(0, weight=1)
    ventana.columnconfigure(2, weight=1)

    # Mostrar elementos de la pantalla de inicio
    frame_label.grid(row=0, column=0, columnspan=6, sticky=W+E)
    home_label.grid(row=0, column=3)
    label_img.grid(row=0, column=2, padx=8, sticky=E)
    contenedor.grid(row=1, column=0, padx=10, pady=10, sticky=W)

    # Mostrar botones y etiquetas en add_frame

    castellar.grid(row=0, column=0, pady=3, padx=2, sticky=W)
    cancelar1.grid(row=0, column=1, pady=3, padx=24)
    castellar_crono.grid(row=0, column=2, pady=3, padx=1)

    gomez.grid(row=1, column=0, pady=3, padx=2)
    cancelar2.grid(row=1, column=1, pady=3, padx=24)
    gomez_crono.grid(row=1, column=2, pady=3, padx=1)

    Castrillo.grid(row=2, column=0, pady=3, padx=2)
    cancelar3.grid(row=2, column=1, pady=3, padx=24)
    Castrillo_crono.grid(row=2, column=2, pady=3, padx=1)

    Virgilio.grid(row=3, column=0, pady=3, padx=2)
    cancelar4.grid(row=3, column=1, pady=3, padx=24)
    virgilio_crono.grid(row=3, column=2,  padx=1)

    gomezJose.grid(row=4, column=0, pady=3, padx=2)
    cancelar5.grid(row=4, column=1, pady=3, padx=24)
    gomez_Jose_crono.grid(row=4, column=2, pady=3, padx=1)

    botones_frm.grid(row=5, column=0, padx=5, pady=20)
    boton_salida.grid(row=1, column=0, padx=10, pady=7)
    btn_ingresar.grid(row=1, column=2, padx=10, pady=7)

    return True

# Funciones para mover las ventanas secundarias

def guardar_posi_secundaria(event):
    global x_origen_sec, y_origen_sec
    x_origen_sec, y_origen_sec = event.x, event.y


def mover_ventana_secundaria(event, ventana_secundaria):
    x, y = ventana_secundaria.winfo_x() + event.x - \
        x_origen_sec, ventana_secundaria.winfo_y() + event.y - y_origen_sec
    if x < 0:
        x = 0
    if y < 0:
        y = 0
    ventana_secundaria.geometry(f"+{x}+{y}")
       


# Función para abrir la ventana ingresar la clave
def ventana_ingresar():
    global ventana_clave
    ventana.withdraw()  # Oculta la ventana principal
    ventana_clave = Toplevel()
    ventana_clave.title("Ingresar")
    ventana_clave.geometry("260x360")
    ventana_clave.resizable(0, 0)
    ventana_clave.overrideredirect(True)
    # Establecer esta ventana secundaria como "siempre arriba"
    ventana_clave.wm_attributes("-topmost", True)
    ventana_clave.configure(bg="gray88")

    x_principal, y_principal = obtener_posicion_ventana_principal()

    # Calcular la nueva posición para la ventana secundaria (justo al lado de la ventana principal)
    # Agrega un espacio de 10 píxeles entre ambas ventanas
    x_secundaria = x_principal + ventana.winfo_width() - 260
    y_secundaria = y_principal

    # Establecer la nueva posición para la ventana secundaria
    ventana_clave.geometry(f"+{x_secundaria}+{y_secundaria}")

    ventana_clave.bind("<Button-1>", guardar_posi_secundaria)
    ventana_clave.bind("<B1-Motion>", lambda event: mover_ventana_secundaria(event, ventana_clave))

    LabelEncabezado = Label(ventana_clave, text="Ingresar BD")
    LabelEncabezado.grid(row=0, column=0, sticky=W)
    LabelEncabezado.config(
        fg="white",
        bg="black",
        font=("Segoe UI", 10, "bold"),
        justify="right",
        padx=95,
        pady=8
    )
    # Desactivar Boton de ingreso
    deshabilitar_botones()
    
    # frame ventana de ingreso
    frm_ingreso = LabelFrame(ventana_clave, bg="gray88", padx=25, pady=10)
    frm_ingreso.grid(row=2, column=0, columnspan=6, padx=10, pady=30)

    # Campos de contraseña
    LabelContraseña = Label(
        frm_ingreso, text="Introduce Contraseña", bg="gray88")
    LabelContraseña.grid(row=3, column=0, pady=2)
    campo_cc_entry = Entry(frm_ingreso, show="*", border=1, width=20,
                           font=("Segoe UI", 11, "bold"), bg="lightBlue", fg="black")
    campo_cc_entry.grid(row=4, column=0)
    campo_cc_entry.focus_set()

    def cerrar_ventana_clave():
        # Activar los campos y botones en la ventana principal
        habilitar_botones()
        ventana.deiconify()  # Muestra la ventana principal

        # Cerrar la ventana secundaria
        ventana_clave.destroy()

    # Funcion para asociar la tecla enter al comando
    def ingresar(event=None):
        verificar_contrasena()

    def verificar_contrasena():
        contrasena_ingresada = campo_cc_entry.get()
        # Contraseña predefinida (puedes cambiarla)
        contrasena_correcta = "123456"

        if contrasena_ingresada == contrasena_correcta:
            ventana_clave.destroy()
            ventana_exportar()

        # Aquí puedes realizar las acciones que deseas después de verificar la contraseña correctamente
        else:
            messagebox.showerror(
                "Verificación de Contraseña", "Contraseña incorrecta")
            campo_cc_entry.delete(0, 'end')  # Limpiar el campo de contraseña
            campo_cc_entry.focus_set()

    frame_btn_ingreso = LabelFrame(ventana_clave, bg="gray88", padx=5, pady=5)
    frame_btn_ingreso.grid(row=5, column=0,  pady=130, columnspan=10)

    # Asociar la función "ingresar" al evento <Return> (tecla Enter) en el campo de entrada de contraseña
    campo_cc_entry.bind("<Return>", ingresar)
    boton_contraseña = Button(
        frame_btn_ingreso, text="Verificar", image=open_img, command=verificar_contrasena)
    boton_contraseña.grid(row=6, column=0, padx=10, sticky=W)
    botones_ingresar(boton_contraseña)

    # boton cerrar en la ventana de contraseña
    boton_cerrar = Button(frame_btn_ingreso,  text="Cerrar",
                          image=Regresar_img, command=cerrar_ventana_clave)
    boton_cerrar.grid(row=6, column=3, padx=10, sticky=E)
    estilo_botones_Salidas(boton_cerrar)
    
    # Asociar la acción de cerrar a la función cerrar_ventana_secundaria
    ventana_clave.protocol("WM_DELETE_WINDOW", cerrar_ventana_clave)


# Establecer la configuración de localización en español colombiano
locale.setlocale(locale.LC_ALL, 'es_CO.utf8')

def regresar_home():
    ventana_expo.destroy()
    habilitar_botones()
    ventana.deiconify()

# Función de la ventana exportar
def ventana_exportar():
    global ventana_expo
    ventana_expo = Toplevel()
    ventana_expo.title("Consulta")
    ventana_expo.geometry("300x350")
    ventana_expo.resizable(0, 0)
    ventana_expo.overrideredirect(True)
    # Establecer esta ventana secundaria como "siempre arriba"
    ventana_expo.wm_attributes("-topmost", True)
    ventana_expo.configure(bg="gray88")

    x_principal, y_principal = obtener_posicion_ventana_principal()

    # Calcular la nueva posición para la ventana secundaria (justo al lado de la ventana principal)
    # Agrega un espacio de 10 píxeles entre ambas ventanas
    x_secundaria = x_principal + ventana.winfo_width() - 260
    y_secundaria = y_principal

    # Establecer la nueva posición para la ventana secundaria
    ventana_expo.geometry(f"+{x_secundaria}+{y_secundaria}")

    ventana_expo.bind("<Button-1>", guardar_posi_secundaria)
    ventana_expo.bind(
        "<B1-Motion>", lambda event: mover_ventana_secundaria(event, ventana_expo))

    LabelEncabezado = Label(ventana_expo, text="Exportar por Fechas")
    LabelEncabezado.grid(row=0, column=0, sticky=W)
    LabelEncabezado.config(
        fg="white",
        bg="black",
        font=("Segoe UI", 10, "bold"),
        justify="right",
        padx=90,
        pady=8)

    style = ttk.Style()
    style.theme_use('clam')

    frameExportar = LabelFrame(ventana_expo, bg="gray88", padx=5, pady=5)
    frameExportar.grid(row=1, column=0, pady=10)

    fecha_inicio_label = Label(
        frameExportar, text="Fecha de inicio", bg="gray88")
    fecha_inicio_label.grid(row=2, column=0, padx=20, sticky=W)

    fecha_inicio_entry = DateEntry(
        frameExportar, locale="es_CO", date_pattern="dd/mm/yyyy")
    fecha_inicio_entry.grid(row=3, column=0, padx=20, sticky=W)

    fecha_fin_label = Label(frameExportar, text="Fecha de final", bg="gray88")
    fecha_fin_label.grid(row=2, column=1, padx=20, sticky=W)

    fecha_fin_entry = DateEntry(frameExportar, date_pattern="dd/mm/yyyy")
    fecha_fin_entry.grid(row=3, column=1, padx=20, sticky=E)

    framebotones = LabelFrame(ventana_expo, bg="gray88", padx=5, pady=5)
    framebotones.grid(row=4, column=0, pady=170, sticky=S)

    def cerrar_ventana_exportar():
        # Activar los campos y botones en la ventana principal
        habilitar_botones()
        
        ventana.deiconify()  # Muestra la ventana principal
        
        # Cerrar la ventana secundaria
        ventana_expo.destroy()

    # Función para exportar datos
    def exportar_datos_excel():
        fecha_inicio = fecha_inicio_entry.get_date()
        fecha_final = fecha_fin_entry.get_date()

        # Convertir las fechas a formato datetime
        fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
        fecha_final = datetime.combine(fecha_final, datetime.max.time())

        # Obtener la fecha siguiente a la seleccionada
        fecha_siguiente = fecha_final + timedelta(days=1)

        try:
            conn = sqlite3.connect(ruta_basedatos)
            cursor = conn.cursor()

            # Consulta a la base de datos utilizando las fechas seleccionadas
            cursor.execute("""
            SELECT r.nombreSup, r.fechahorainicio, r.fechahorafinal, r.fechahorafinal - r.fechahorainicio AS Tiempo_Reaccion, 
                        r.observacion, r.nombre_terminal, t.nombre
                        FROM Reaccion r
                        INNER JOIN turno t ON r.idturno = t.ID
                        WHERE r.fechahorainicio >= ? AND r.fechahorafinal < ?
                        AND r.observacion IS NULL  -- Filtrar observaciones nulas
                        ORDER BY r.fechahorafinal DESC
                        """, (fecha_inicio, fecha_siguiente))
            result = cursor.fetchall()

            if result:
                # Crear un DataFrame de Pandas con los resultados de la consulta
                df = pd.DataFrame(result, columns=[
                                  "Nombre_Supervisor", "fecha_inicio", "fecha_final", "Tiempo_Reaccion", "Observacion", "nombre_terminal", "Turno"])

                # Convertir las columnas de fecha al formato de objeto de fecha
                df['fecha_inicio'] = pd.to_datetime(
                    df['fecha_inicio'], format='%Y-%m-%d %H:%M:%S')
                df['fecha_final'] = pd.to_datetime(
                    df['fecha_final'], format='%Y-%m-%d %H:%M:%S')

                # Calcular la diferencia de tiempo (duración) entre fecha_final y fecha_inicio
                df['Tiempo_Reaccion'] = df['fecha_final'] - df['fecha_inicio']

                # Formatear la columna Tiempo_Reaccion en el formato HH:MM:SS
                df['Tiempo_Reaccion'] = df['Tiempo_Reaccion'].apply(
                    lambda x: str(x).split()[-1])

                # Crear un archivo Excel y agregar hojas para cada nombre de supervisor
                file_path = "datos_reaccion.xlsx"
                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                    for nombre_supervisor, datos_supervisor in df.groupby("Nombre_Supervisor"):
                        datos_supervisor.to_excel(
                            writer, sheet_name=nombre_supervisor, index=False)

                    # Calcular el promedio semanal de Tiempo_Reaccion para todos los supervisores
                    df['Semana'] = df['fecha_inicio'].dt.strftime(
                        '%Y-%U')  # Año-Semana
                    df['Tiempo_Reaccion'] = pd.to_timedelta(
                        df['Tiempo_Reaccion'])  # Convertir a timedelta

                    # Crear una tabla pivote para mostrar los promedios semanales por supervisor
                    promedio_pivot = df.pivot_table(
                        index='Nombre_Supervisor', columns='Semana', values='Tiempo_Reaccion', aggfunc='mean')

                    # Obtener la lista completa de semanas
                    semanas_completas = pd.date_range(start=df['fecha_inicio'].min(
                    ), end=df['fecha_inicio'].max(), freq='W-MON').strftime('%Y-%U')

                    # Agregar las semanas faltantes en el DataFrame
                    promedio_pivot = promedio_pivot.reindex(
                        columns=semanas_completas)

                    # Llenar los valores NaN (que corresponden a semanas sin datos) con 0
                    promedio_pivot = promedio_pivot.fillna(
                        pd.Timedelta(seconds=0))

                    # Calcular el promedio
                    promedio_pivot['TOTAL_PROMEDIO'] = promedio_pivot.mean(
                        axis=1)

                    # Formatear los valores timedelta en formato HH:MM:SS sin milisegundos
                    promedio_pivot = promedio_pivot.applymap(
                        lambda x: str(x).split()[-1].split('.')[0])

                    # Crear la hoja de promedios en el archivo Excel
                    promedio_pivot.to_excel(
                        writer, sheet_name='Promedio_Semanal', index=True)

                    # Obtener la hoja de promedios recién creada
                    hoja_promedios = writer.sheets['Promedio_Semanal']

                    # Recorrer las filas y resaltar las celdas en amarillo cuando el valor sea NaN
                    for fila in hoja_promedios.iter_rows(min_row=2, max_row=hoja_promedios.max_row, min_col=2, max_col=hoja_promedios.max_column):
                        for celda in fila:
                            if celda.value == '00:00:00':
                                celda.fill = PatternFill(
                                    start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

                                # Cambiar el color de fuente a blanco
                                celda.font = Font(
                                    color="FFFFFF", bold=True, underline="single")

                messagebox.showinfo(
                    "Registros encontrados", f"Se encontraron {len(result)} registros para la fecha seleccionada")
                
                regresar_home()  #regresar a la ventana principal
                
            else:
                messagebox.showinfo(
                    "Sin resultados", "No se encontraron datos para las fechas seleccionadas")
            conn.close()

        except sqlite3.Error as e:
            print("Error al exportar los datos:", e)

    boton_buscar = Button(framebotones, text="Exportar",
                          image=Export_img,  command=exportar_datos_excel)
    boton_buscar.grid(row=5, column=0, pady=5, padx=20)
    botones_ingresar(boton_buscar)

    # boton salir de la ventana exportar
    boton_Salir = Button(framebotones, text="Cerrar",
                         image=Regresar_img, command=cerrar_ventana_exportar)
    boton_Salir.grid(row=5, column=1, pady=5, padx=20)
    estilo_botones_Salidas(boton_Salir)

    # Configurar el evento de cierre de la ventana secundaria
    ventana_expo.protocol("WM_DELETE_WINDOW", cerrar_ventana_exportar)


frame_label = Frame(ventana, bg="black")
# Definir campos de pantallas (Inicio)
home_label = Label(frame_label, text="Reacción de Supervisores", anchor=W)
label_img = Label(frame_label, image=Reloj_img, bg="black", anchor=E)
contenedor = LabelFrame(ventana, padx=8, pady=5, border=2, bg="gray88")

# funcion que llama a las funciones eliminar tabla temporal y cancelar cronometro


def eliminar_tabla_temp(label, boton, cancelar):
    eliminar_tabla_temporal(boton["text"])
    cancelar_cronometro(label, boton, cancelar)

# funcion que llama a las funciones eliminar tabla temporal y cancelar cronometro cuando se ha cerrado la aplicacion


def eliminar_tabla(label, boton, cancelar):
    eliminar_tabla_t(boton["text"])
    cancelar_cronometro(label, boton, cancelar)
    boton.config(state="normal")

# FUNCION PARA INICIAR EL CANCELAR ELIMINAR TABLA


def iniciar_proceso(boton, label, cancelar):
    cancelar.config(state='normal')
    boton["bg"] = "red"
    boton.config(state="disabled")
    boton["fg"] = "white"
    cancelar.bind("<Button-1>", lambda event: eliminar_tabla(label,
                  boton, cancelar) if cancelar["state"] == "normal" else None)

# FUNCION PARA CAMBIAR COLOR DEL BOTON E INICIAR TEMPORIZADOR


def cambiar_color_y_temporizador(boton, label, cancelar):
    # Obtener el nombre del botón
    nombre_boton = boton["text"]
    # Remover caracteres no válidos del nombre del botón
    nombre_boton = nombre_boton.replace(" ", "_")

    if boton["bg"] == "green2":
        # Nombre de la tabla temporal
        nombre_tabla = f"temp_reaccion_{nombre_boton}"

        # Conexión a la base de datos SQLite compartida
        conn = sqlite3.connect(ruta_basedatos)
        cursor = conn.cursor()

        # Verificar si la tabla ya existe
        cursor.execute(
            f"SELECT name FROM sqlite_master WHERE type='table' AND name='{nombre_tabla}'")
        tabla_existente = cursor.fetchone()

        # Si la tabla ya existe, mostrar mensaje y salir de la función
        if tabla_existente:
            # Obtener el nombre de la terminal almacenado en la tabla temporal
            cursor.execute(
                f"SELECT nombre_terminal FROM {nombre_tabla} LIMIT 1")
            nombre_terminal_tabla = cursor.fetchone()
            nombre_terminal = nombre_terminal_tabla[0]
            # Mostrar mensaje informativo
            messagebox.showinfo(
                "Información", f"Existe un Proceso Iniciado para el Supervisor desde la Terminal de {nombre_terminal}.")
            iniciar_proceso(boton, label, cancelar)
            conn.close()
            return

        boton["bg"] = "red"
        cancelar.config(state='normal')
        boton["fg"] = "white"
        label["text"] = "00:00"
        label["fg"] = "black"
        label["font"] = ("Arial", 12, "bold")
        cancelar.bind("<Button-1>", lambda event: eliminar_tabla_temp(label,
                      boton, cancelar) if cancelar["state"] == "normal" else None)
        start_time = time.time()
        actualizar_temporizador(label, start_time, boton)

        # Llamada a la función para crear la tabla temporal
        crear_tabla_temporal(nombre_boton)
        conn.close()
    else:
        cancelar.config(state='disabled')
        boton["bg"] = "green2"
        boton["fg"] = "black"
        label["text"] = "00:00"
        label["fg"] = "black"
        label["font"] = ("Arial", 12, "bold")
        cancelar.unbind("<Button-1>")
        boton.reset = True

        # Llamada a la función para actualizar y guardar los datos
        if boton.reset:
            actualizar_y_guardar_datos(nombre_boton)

 # funcion para obtener el Id del turno


def obtener_id_turno(cursor, hora_actual):
    cursor.execute("SELECT ID FROM turno WHERE horainicio <= horafinal AND horainicio <= ? AND horafinal >= ?",
                   (hora_actual, hora_actual))
    resultado = cursor.fetchone()
    if resultado:
        return resultado[0]
    else:
        cursor.execute("SELECT ID FROM turno WHERE horainicio > horafinal AND (horainicio <= ? OR horafinal >= ?)",
                       (hora_actual, hora_actual))
        resultado = cursor.fetchone()
        if resultado:
            return resultado[0]
        else:
            return None

# funcion para crear la tabla de los supervisores


def crear_tabla_temporal(nombre_boton):
    try:
        # Remover caracteres no válidos del nombre del botón
        nombre_boton = nombre_boton.replace(" ", "_")

        # Nombre de la tabla temporal
        nombre_tabla = f"temp_reaccion_{nombre_boton}"

        # Conexión a la base de datos SQLite compartida
        conn = sqlite3.connect(ruta_basedatos)
        cursor = conn.cursor()

        # Crear tabla temporal si no existe
        cursor.execute(f"""CREATE TABLE IF NOT EXISTS {nombre_tabla} (Nombre_sup TEXT, fecha_hora_inicio DATETIME, fecha_hora_final DATETIME, 
                       MinutSegundos TEXT, observacion TEXT, nombre_terminal TEXT, idturno INT)""")

        # Obtener la fecha y hora actual
        fecha_hora_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Obtener la fecha y hora actual
        hora_actual = datetime.now().strftime('%H:%M')

        # Obtener el ID del turno basado en la hora actual
        id_turno = obtener_id_turno(cursor, hora_actual)

        if id_turno is not None:
            # Insertar registro en la tabla temporal con el ID del turno
            cursor.execute(f"INSERT INTO {nombre_tabla} (Nombre_sup, fecha_hora_inicio, nombre_terminal, idturno) VALUES (?, ?, ?, ?)", (
                nombre_boton, fecha_hora_actual, nombre_terminal, id_turno))
            # Guardar cambios (commit) sin cerrar la conexión
            conn.commit()
        else:
            print("No se encontró el turno asociado al botón:", nombre_boton)
    except sqlite3.Error as e:
        print("Error en la operación de base de datos:", e)
    finally:
        conn.close()

# Función para actualizar datos de la tabla temporal e insertar en la tabla Reaccion


def actualizar_y_guardar_datos(nombre_boton):
    # Remover caracteres no válidos del nombre del botón
    nombre_boton = nombre_boton.replace(" ", "_")

    # Nombre de la tabla temporal
    nombre_tabla_temporal = f"temp_reaccion_{nombre_boton}"

    # Conexión a la base de datos SQLite compartida
    conn = sqlite3.connect(ruta_basedatos)
    cursor = conn.cursor()

    # Obtener la fecha y hora actual
    fecha_hora_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Obtener el tiempo transcurrido de la variable global
    minutos_segundos = tiempo_transcurrido

    horas = int(tiempo_transcurrido // 3600)
    minutos = int((tiempo_transcurrido % 3600) // 60)
    segundos = int(tiempo_transcurrido % 60)

    # Formatear la diferencia de tiempo en HH:mm:ss
    minutos_segundos = f"{horas:02d}:{minutos:02d}:{segundos:02d}"

    # Actualizar los campos en la tabla temporal con el nombre de la terminal desde el archivo config
    cursor.execute(f"UPDATE {nombre_tabla_temporal} SET fecha_hora_final = ?, MinutSegundos = ? WHERE Nombre_sup = ?",
                   (fecha_hora_actual, minutos_segundos, nombre_boton))

    # Obtener los datos actualizados de la tabla temporal
    cursor.execute(
        f"SELECT * FROM {nombre_tabla_temporal} WHERE Nombre_sup = ?", (nombre_boton,))
    datos_temporales = cursor.fetchone()

    # Insertar los datos en la tabla Reaccion
    cursor.execute("""INSERT INTO Reaccion (nombreSup, fechahorainicio, fechahorafinal, MinutSegundos, observacion, nombre_terminal, 
                   idturno) VALUES (?, ?, ?, ?, ?, ?, ?)""", (datos_temporales[0], datos_temporales[1], datos_temporales[2], datos_temporales[3],
                                                              datos_temporales[4], datos_temporales[5], datos_temporales[6]))

    # Eliminar la tabla temporal
    cursor.execute(f"DROP TABLE IF EXISTS {nombre_tabla_temporal}")

    # Guardar cambios (commit) y cerrar la conexión
    conn.commit()
    conn.close()


# Leer la configuración desde el archivo .ini
config = configparser.ConfigParser()
config.read('config.ini')

# Obtener la ruta de la base de datos y el nombre de la terminal desde la configuración
ruta_basedatos = config.get('Configuracion', 'ruta_basedatos')
nombre_terminal_config = config.get('Configuracion', 'nombre_terminal')

# Función para eliminar la tabla temporal asociada a un botón


def eliminar_tabla_temporal(nombre_boton):
    # Remover caracteres no válidos del nombre del botón
    nombre_boton = nombre_boton.replace(" ", "_")

    # Nombre de la tabla temporal
    nombre_tabla_temporal = f"temp_reaccion_{nombre_boton}"

    # Conexión a la base de datos SQLite compartida
    conn = sqlite3.connect(ruta_basedatos)
    cursor = conn.cursor()

    # Obtener la fecha y hora actual
    fecha_hora_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Obtener el tiempo transcurrido de la variable global (asegúrate de que esta variable esté definida)
    minutos_segundos = tiempo_transcurrido

    # Calcular los minutos y segundos a partir del tiempo transcurrido en segundos
    minutos = tiempo_transcurrido // 60
    segundos = tiempo_transcurrido % 60

    # Formatear los minutos y segundos en una cadena "MM:SS"
    minutos_segundos = f"{minutos:02d}:{segundos:02d}"

    # Obtener el nombre de la terminal almacenado en la tabla temporal
    cursor.execute(
        f"SELECT nombre_terminal FROM {nombre_tabla_temporal} LIMIT 1")
    nombre_terminal_tabla = cursor.fetchone()

    if nombre_terminal_tabla and nombre_terminal_tabla[0] == nombre_terminal_config:
        # Actualizar los campos en la tabla temporal con el nombre de la terminal desde el archivo config
        cursor.execute(f"UPDATE {nombre_tabla_temporal} SET fecha_hora_final = ?, MinutSegundos = ?, observacion = ?  WHERE Nombre_sup = ?", (
            fecha_hora_actual, minutos_segundos, 'Cancelado', nombre_boton))

        # Obtener los datos actualizados de la tabla temporal
        cursor.execute(
            f"SELECT * FROM {nombre_tabla_temporal} WHERE Nombre_sup = ?", (nombre_boton,))
        datos_temporales = cursor.fetchone()

        # Insertar los datos en la tabla Reaccion
        cursor.execute("""INSERT INTO Reaccion (nombreSup, fechahorainicio, fechahorafinal, MinutSegundos, observacion, 
                       nombre_terminal, idturno) VALUES (?, ?, ?, ?, ?, ?, ?)""", (datos_temporales[0], datos_temporales[1], datos_temporales[2], datos_temporales[3], datos_temporales[4],
                                                                                   datos_temporales[5], datos_temporales[6]))

        # Eliminar la tabla temporal
        cursor.execute(f"DROP TABLE IF EXISTS {nombre_tabla_temporal}")
        # Guardar cambios (commit) y cerrar la conexión
        conn.commit()

    else:
        mensaje = f"Cancelar desde la terminal de {nombre_terminal_tabla[0]}"
        messagebox.showinfo("Informacion", mensaje)
        conn.close()

# Función para eliminar la tabla temporal asociada a un botón


def eliminar_tabla_t(nombre_boton):
    # Remover caracteres no válidos del nombre del botón
    nombre_boton = nombre_boton.replace(" ", "_")

    # Nombre de la tabla temporal
    nombre_tabla_temporal = f"temp_reaccion_{nombre_boton}"

    # Conexión a la base de datos SQLite compartida
    conn = sqlite3.connect(ruta_basedatos)
    cursor = conn.cursor()

    # Obtener la fecha y hora actual en formato de texto
    fecha_hora_actual_texto = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Obtener el nombre de la terminal almacenado en la tabla temporal
    cursor.execute(
        f"SELECT nombre_terminal, fecha_hora_inicio FROM {nombre_tabla_temporal} LIMIT 1")
    nombre_terminal_tabla, fecha_hora_inicio_texto = cursor.fetchone()

    if nombre_terminal_tabla and nombre_terminal_tabla == nombre_terminal_config:
        # Convertir las cadenas de texto en objetos datetime
        fecha_hora_actual = datetime.strptime(
            fecha_hora_actual_texto, '%Y-%m-%d %H:%M:%S')
        fecha_hora_inicio = datetime.strptime(
            fecha_hora_inicio_texto, '%Y-%m-%d %H:%M:%S')

        # Calcular la diferencia de tiempo
        diferencia_tiempo = fecha_hora_actual - fecha_hora_inicio

        # Calcular horas, minutos y segundos
        segundos_totales = diferencia_tiempo.total_seconds()
        horas = int(segundos_totales // 3600)
        minutos = int((segundos_totales % 3600) // 60)
        segundos = int(segundos_totales % 60)

        # Formatear la diferencia de tiempo en HH:mm:ss
        minutos_segundos = f"{horas:02d}:{minutos:02d}:{segundos:02d}"

        # Actualizar los campos en la tabla temporal con la diferencia de tiempo calculada
        cursor.execute(f"UPDATE {nombre_tabla_temporal} SET fecha_hora_final = ?, MinutSegundos = ?, observacion = ?  WHERE Nombre_sup = ?", (
            fecha_hora_actual_texto, minutos_segundos, 'Cancelado por Cierre del Sistema', nombre_boton))

        # Obtener los datos actualizados de la tabla temporal
        cursor.execute(
            f"SELECT * FROM {nombre_tabla_temporal} WHERE Nombre_sup = ?", (nombre_boton,))
        datos_temporales = cursor.fetchone()

        # Insertar los datos en la tabla Reaccion
        cursor.execute("""INSERT INTO Reaccion (nombreSup, fechahorainicio, fechahorafinal, MinutSegundos, observacion, 
                       nombre_terminal, idturno) VALUES (?, ?, ?, ?, ?, ?, ?)""", (datos_temporales[0], datos_temporales[1], datos_temporales[2], datos_temporales[3], datos_temporales[4],
                                                                                   datos_temporales[5], datos_temporales[6]))

        # Eliminar la tabla temporal
        cursor.execute(f"DROP TABLE IF EXISTS {nombre_tabla_temporal}")
        # Guardar cambios (commit) y cerrar la conexión
        conn.commit()
    else:
        mensaje = f"Cancelar desde la terminal de {nombre_terminal_tabla}."
        messagebox.showinfo("Informacion", mensaje)
        conn.close()


# Declarar una variable global para almacenar el tiempo transcurrido
tiempo_transcurrido = 0

# Función para actualizar el temporizador cada segundo


def actualizar_temporizador(label, start_time, boton):
    global tiempo_transcurrido  # Acceder a la variable global
    elapsed_time = time.time() - start_time
    minutos = int(elapsed_time // 60)
    segundos = int(elapsed_time % 60)
    # Actualizar la variable global con el tiempo transcurrido
    tiempo_transcurrido = minutos * 60 + segundos

    if boton.reset:
        label["text"] = "00:00"
        label["fg"] = "black"
        boton.reset = False
    else:
        label["text"] = f"{minutos:02d}:{segundos:02d}"
    if minutos >= 10:
        animar_label(label)
    elif minutos >= 5:
        ani_label_5min(label)
    if boton["bg"] == "red":
        ventana.after(1000, actualizar_temporizador, label, start_time, boton)

# Función para cancelar el cronómetro


def cancelar_cronometro(label, boton, cancelar):
    boton["bg"] = "green2"
    boton["fg"] = "black"
    label["text"] = "00:00"
    label["font"] = ("Arial", 12, "bold")
    cancelar.config(state='disabled')
    cancelar.unbind("<Button-1>")
    boton.reset = True


# frame botones
botones_frm = LabelFrame(ventana, padx=7, pady=5, border=2, bg="gray88")

# Funcion para verificar si existe algun proceso iniciado por el supervisor


def verificar_proceso_supervisor(botones):
    # Verificar si algún botón está en rojo
    for boton in botones:
        if boton["bg"] == "red":
            messagebox.showinfo(
                "Proceso del Supervisor", "Termine el proceso activo antes de cerrar la ventana.")
            return
    ventana.destroy()


# Crear botón Castellar, cancelar y cronómetro
castellar = Button(contenedor, text="Castellar", command=lambda: cambiar_color_y_temporizador(
    castellar, castellar_crono,  cancelar1))
estilos_botones_supervisores(castellar)
castellar.reset = False
cancelar1 = Label(contenedor, state="disabled", bg="gray88")
det_cancelar(cancelar1)
castellar_crono = Label(contenedor, text="00:00", fg="black", font=(
    "Arial", 12, "bold"), bg="gray88", padx=1)

# Crear botón Gómez, cancelar y cronómetro
gomez = Button(contenedor, text="Gómez", command=lambda: cambiar_color_y_temporizador(
    gomez, gomez_crono,  cancelar2))
estilos_botones_supervisores(gomez)
gomez.reset = False
cancelar2 = Label(contenedor, state="disabled", bg="gray88")
det_cancelar(cancelar2)
gomez_crono = Label(contenedor, text="00:00", fg="black",
                    font=("Arial", 12, "bold"), bg="gray88", padx=1)

# Crear botón Gutierrez, cancelar y cronómetro
Castrillo = Button(contenedor, text="Castrillo", command=lambda: cambiar_color_y_temporizador(
    Castrillo, Castrillo_crono,  cancelar3))
estilos_botones_supervisores(Castrillo)
Castrillo.reset = False
cancelar3 = Label(contenedor, state="disabled", bg="gray88")
det_cancelar(cancelar3)
Castrillo_crono = Label(contenedor, text="00:00", fg="black", font=(
    "Arial", 12, "bold"), bg="gray88", padx=1)

# Crear botón Virgilio, cancelar y cronómetro
Virgilio = Button(contenedor, text="Virgilio", command=lambda: cambiar_color_y_temporizador(
    Virgilio, virgilio_crono,  cancelar4))
estilos_botones_supervisores(Virgilio)
Virgilio.reset = False
cancelar4 = Label(contenedor, state="disabled", bg="gray88")
det_cancelar(cancelar4)
virgilio_crono = Label(contenedor, text="00:00", fg="black", font=(
    "Arial", 12, "bold"), bg="gray88", padx=1)

# Crear botón Gomez Jose, cancelar y cronómetro
gomezJose = Button(contenedor, text="Gomez J", command=lambda: cambiar_color_y_temporizador(
    gomezJose, gomez_Jose_crono,  cancelar5))
estilos_botones_supervisores(gomezJose)
gomezJose.reset = False
cancelar5 = Label(contenedor, state="disabled", bg="gray88")
det_cancelar(cancelar5)
gomez_Jose_crono = Label(contenedor, text="00:00", fg="black", font=(
    "Arial", 12, "bold"), bg="gray88", padx=1)

# Definir la lista de botones
botones = [castellar, gomez, Castrillo, Virgilio, gomezJose]

# Crear botón de salida
boton_salida = Button(botones_frm, text="Salir", image=Salir_img,
                      command=lambda: verificar_proceso_supervisor(botones))
estilo_botones_Salidas(boton_salida)

# crear boton ingresar
btn_ingresar = Button(botones_frm, text="Ingresar",
                      image=Ingresar_img, command=ventana_ingresar)
botones_ingresar(btn_ingresar)

# cargar pantalla inicio
home()

# Iniciar el bucle de eventos
ventana.mainloop()
